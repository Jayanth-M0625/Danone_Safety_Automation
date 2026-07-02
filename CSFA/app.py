import os
import json
import pandas as pd
import numpy as np
import openpyxl
import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
from trend_classifier_ml import predict_trend

# Base directory of this script to resolve relative paths correctly
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def get_path(filename):
    return os.path.join(BASE_DIR, filename)

# Page config (Commented out because this is run as a sub-page under st.navigation)
# st.set_page_config(
#     page_title="Contractor Safety Field Audits Dashboard - Danone",
#     page_icon="🛡️",
#     layout="wide",
#     initial_sidebar_state="expanded"
# )

# Custom premium CSS injection
st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;800&display=swap" rel="stylesheet">
<style>
    /* Global styles */
    html, body, [class*="css"] {
        font-family: 'Outfit', sans-serif !important;
    }
    
    .main .block-container {
        padding-top: 1.5rem;
        padding-bottom: 1.5rem;
    }
    
    /* Header section styling */
    .dashboard-header {
        background: linear-gradient(135deg, #002060 0%, #0033a0 100%);
        padding: 20px 25px;
        border-radius: 10px;
        color: white;
        margin-bottom: 20px;
        box-shadow: 0 4px 15px rgba(0, 51, 160, 0.15);
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .header-title-box {
        display: flex;
        flex-direction: column;
    }
    .header-main-title {
        font-size: 24px;
        font-weight: 800;
        letter-spacing: 0.5px;
        margin: 0;
        text-transform: uppercase;
    }
    .header-sub-title {
        font-size: 13px;
        font-weight: 400;
        opacity: 0.85;
        margin-top: 2px;
    }
    .header-info-box {
        text-align: right;
        background: rgba(255, 255, 255, 0.1);
        padding: 8px 12px;
        border-radius: 6px;
        backdrop-filter: blur(5px);
        border: 1px solid rgba(255, 255, 255, 0.2);
    }
    .header-info-date {
        font-size: 12px;
        font-weight: 600;
    }
    .header-info-lbl {
        font-size: 10px;
        opacity: 0.75;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    /* Target Streamlit's native border container for styling */
    div[data-testid="stVerticalBlockBorderDiv"] {
        background-color: white !important;
        border-radius: 10px !important;
        border: 1px solid #e9ecef !important;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.03) !important;
        padding: 18px !important;
    }

    /* Sidebar styles */
    [data-testid="stSidebar"] {
        background-color: #001233;
        color: white;
    }
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] {
        color: white;
    }
    [data-testid="stSidebar"] button {
        background-color: #0033a0 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
    }
    [data-testid="stSidebar"] button:hover {
        background-color: #002060 !important;
        box-shadow: 0 4px 10px rgba(0, 0, 0, 0.3) !important;
    }
</style>
""", unsafe_allow_html=True)

# ----------------- DATA PIPELINE FUNCTIONS -----------------

def normalize_text(text):
    if pd.isna(text):
        return ""
    return "".join(c.lower() for c in str(text) if c.isalnum())

def run_sync(drcsfa_path=None, csfa_path=None, metadata_path=None):
    if drcsfa_path is None:
        drcsfa_path = get_path("DRCSFA.xlsx")
    if csfa_path is None:
        csfa_path = get_path("CSFA Accumilative data.xlsx")
    if metadata_path is None:
        metadata_path = get_path("sync_metadata.json")

    if os.path.exists(metadata_path):
        with open(metadata_path, 'r') as f:
            meta = json.load(f)
    else:
        meta = {"last_processed_drcsfa_id": 216}
        
    last_id = meta.get("last_processed_drcsfa_id", 216)
    
    if not os.path.exists(drcsfa_path):
        return -1, f"Error: Source file '{drcsfa_path}' not found."
    
    df_source = pd.read_excel(drcsfa_path)
    new_rows_df = df_source[df_source['ID'] > last_id]
    if new_rows_df.empty:
        return 0, "Already up to date. No new observations found in DRCSFA."
        
    if not os.path.exists(csfa_path):
        return -1, f"Error: Accumulative sheet '{csfa_path}' not found."
        
    wb = openpyxl.load_workbook(csfa_path, data_only=False)
    ws = wb['CSFA Accumilative data']
    
    target_row = 2
    for r in range(2, ws.max_row + 1000):
        val = ws.cell(row=r, column=3).value
        if val is None or str(val).strip() == "":
            target_row = r
            break
            
    start_ln = ws.cell(row=target_row, column=1).value
    if not isinstance(start_ln, (int, float)):
        prev_val = ws.cell(row=target_row - 1, column=1).value
        if isinstance(prev_val, (int, float)):
            start_ln = int(prev_val) + 1
        else:
            start_ln = 1
            
    obs_groups = [
        ('Contractor\xa0', 'Condition', 'Type of Violation', 'Severity Score', 'Violation Description', 'No of Good Citizen', 'No of Violators'),
        ('Contractor\xa02', 'Condition2', 'Type of Violation2', 'Severity Score2', 'Violation Description2', 'No of Good Citizen2', 'No of Violators2'),
        ('Contractor\xa03', 'Condition3', 'Type of Violation3', 'Severity Score3', 'Violation Description3', 'No of Good Citizen3', 'No of Violators3'),
        ('Contractor\xa04', 'Condition4', 'Type of Violation4', 'Severity Score4', 'Violation Description4', 'No of Good Citizen4', 'No of Violators4'),
        ('Contractor\xa05', 'Condition5', 'Type of Violation5', 'Severity Score5', 'Violation Description5', 'No of Good Citizen5', 'No of Violators5'),
        ('Contractor\xa06', 'Condition6', 'Type of Violation6', 'Severity Score6', 'Violation Description6', 'No of Good Citizen6', 'No of Violators6'),
        ('Contractor\xa07', 'Condition7', 'Type of Violation7', 'Severity Score7', 'Violation Description7', 'No of Good Citizen7', 'No of Violators7'),
        ('Contractor\xa08', 'Condition8', 'Type of Violation8', 'Severity Score8', 'Violation Description8', 'No of Good Citizen8', 'No of Violators8'),
        ('Contractor\xa09', 'Condition9', 'Type of Violation9', 'Severity Score9', 'Violation Description9', 'No of Good Citizen9', 'No of Violators9'),
        ('Contractor\xa010', 'Condition10', 'Type of Violation10', 'Severity Score10', 'Violation Description10', 'No of Good Citizen10', 'No of Violators10'),
    ]
    
    appended_count = 0
    new_rows_df = new_rows_df.sort_values('ID')
    
    for idx, row in new_rows_df.iterrows():
        raw_date = row.get('Date of Audit')
        if pd.isna(raw_date):
            raw_date = row.get('Start time')
            
        dt = pd.to_datetime(raw_date)
        date_val = dt.to_pydatetime() if pd.notna(raw_date) else None
        month_name = dt.strftime('%B') if pd.notna(raw_date) else ""
        
        zone = row.get('Zone', '')
        
        for g_idx, group in enumerate(obs_groups, 1):
            col_contractor, col_condition, col_violation_type, col_severity, col_desc, col_good, col_violators = group
            desc_val = row.get(col_desc)
            
            if pd.isna(desc_val) or str(desc_val).strip() == "":
                continue
                
            contractor_val = row.get(col_contractor, '')
            condition_val = str(row.get(col_condition, ''))
            violation_type_val = str(row.get(col_violation_type, ''))
            
            raw_severity = str(row.get(col_severity, '0')).strip()
            if "-" in raw_severity:
                severity_val = raw_severity.split("-")[0].strip()
            else:
                severity_val = raw_severity
                
            try:
                severity_val = int(float(severity_val))
            except ValueError:
                severity_val = 0
                
            good_val = row.get(col_good, 0)
            violators_val = row.get(col_violators, 0)
            
            unsafe_cond = 1 if "unsafe condition" in condition_val.lower() else ""
            unsafe_act = 1 if "unsafe act" in condition_val.lower() else ""
            ppe_non_compliance = 1 if "ppe" in violation_type_val.lower() else ""
            four_five_indicator = 1 if severity_val in [4, 5] else ""
            num_violations = 1
            vx_severity = severity_val
            
            ws.cell(row=target_row, column=1, value=start_ln)
            ws.cell(row=target_row, column=2, value=predict_trend(desc_val))
            ws.cell(row=target_row, column=3, value=str(desc_val).strip())
            ws.cell(row=target_row, column=4, value="")
            ws.cell(row=target_row, column=5, value="")
            ws.cell(row=target_row, column=6, value=good_val if pd.notna(good_val) else 0)
            ws.cell(row=target_row, column=7, value=violators_val if pd.notna(violators_val) else 0)
            ws.cell(row=target_row, column=8, value=num_violations)
            ws.cell(row=target_row, column=9, value=severity_val)
            ws.cell(row=target_row, column=10, value=vx_severity)
            ws.cell(row=target_row, column=11, value=four_five_indicator)
            ws.cell(row=target_row, column=12, value=ppe_non_compliance)
            ws.cell(row=target_row, column=13, value=unsafe_cond)
            ws.cell(row=target_row, column=14, value=unsafe_act)
            ws.cell(row=target_row, column=15, value=str(contractor_val).strip())
            ws.cell(row=target_row, column=16, value="Pending")
            ws.cell(row=target_row, column=17, value="")
            ws.cell(row=target_row, column=18, value=str(zone).strip())
            if date_val:
                ws.cell(row=target_row, column=19, value=date_val)
            else:
                ws.cell(row=target_row, column=19, value="")
            ws.cell(row=target_row, column=20, value=month_name)
            
            start_ln = int(start_ln) + 1
            target_row += 1
            appended_count += 1
            
    wb.save(csfa_path)
    meta["last_processed_drcsfa_id"] = int(new_rows_df['ID'].max())
    with open(metadata_path, 'w') as f:
        json.dump(meta, f)
        
    return appended_count, f"Sync successful! Appended {appended_count} new observations from DRCSFA."

# ----------------- LOAD & CLEAN ACCUMULATIVE DATA -----------------

@st.cache_data(ttl=60)
def load_data(csfa_path=None):
    if csfa_path is None:
        csfa_path = get_path("CSFA Accumilative data.xlsx")

    if not os.path.exists(csfa_path):
        return pd.DataFrame()
        
    df = pd.read_excel(csfa_path, sheet_name="CSFA Accumilative data")
    df = df[df['Observation Discription '].notna()]
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Severity'] = pd.to_numeric(df['Severity '], errors='coerce').fillna(0).astype(int)
    df['unsafe acts'] = pd.to_numeric(df['unsafe acts'], errors='coerce').fillna(0).astype(int)
    df['unsafe Condes'] = pd.to_numeric(df['unsafe Condes'], errors='coerce').fillna(0).astype(int)
    df['Status'] = df['Status'].astype(str).str.strip()
    df['Contractor name'] = df['Contractor name'].astype(str).str.strip()
    df['Trend'] = df['Trend'].astype(str).str.strip()
    df['Zone'] = df['Zone'].astype(str).str.strip()
    return df

# Initialize Data
df_raw = load_data()

# ----------------- SIDEBAR -----------------

st.sidebar.image(get_path("danone_logo.png"), use_container_width=True)
st.sidebar.markdown("<h2 style='text-align: center; color: white; margin-top: 10px;'>⚙️ PIPELINE</h2>", unsafe_allow_html=True)

# Sync Button
if st.sidebar.button("🔄 Sync DRCSFA Form Data", use_container_width=True):
    with st.spinner("Syncing latest observations..."):
        count, msg = run_sync()
        if count >= 0:
            st.sidebar.success(msg)
            st.cache_data.clear()
            df_raw = load_data()
            st.rerun()
        else:
            st.sidebar.error(msg)

st.sidebar.markdown("<hr style='border-color: rgba(255,255,255,0.1);'>", unsafe_allow_html=True)
st.sidebar.markdown("<h4 style='color: white;'>🔍 FILTERS</h4>", unsafe_allow_html=True)

# Date Picker Filter
if not df_raw.empty:
    min_date = df_raw['Date'].min().date() if pd.notna(df_raw['Date'].min()) else datetime.now().date() - timedelta(days=90)
    max_date = df_raw['Date'].max().date() if pd.notna(df_raw['Date'].max()) else datetime.now().date()
    
    start_date = st.sidebar.date_input("Start Date", min_date, min_value=min_date, max_value=max_date)
    end_date = st.sidebar.date_input("End Date", max_date, min_value=min_date, max_value=max_date)
    
    if start_date > end_date:
        st.sidebar.error("Error: Start Date must be before or equal to End Date.")
else:
    start_date = datetime.now().date() - timedelta(days=90)
    end_date = datetime.now().date()

# Manpower input
manpower = st.sidebar.number_input("Average Manpower / Day", min_value=1, value=168, step=1)

# List of invalid contractors to ignore in contractor KPIs
ignore_contractors = [
    'nan', 'no one working at roof area.', 'no one working at the time of visit',
    'no one outside', 'no one was working here', 'common area', 'general', 'all',
    'no work happening here', 'test', 'spot', 'leakage', 'general observations',
    'oil storage shed', 'common for all', 'oil injection area', 'none'
]

# Apply Filters
if not df_raw.empty and start_date <= end_date:
    df_filtered = df_raw[(df_raw['Date'].dt.date >= start_date) & (df_raw['Date'].dt.date <= end_date)]
else:
    df_filtered = pd.DataFrame()

# ----------------- HEADER PANEL -----------------

if not df_filtered.empty:
    date_display_str = f"Range: {start_date.strftime('%d-%b-%Y')} to {end_date.strftime('%d-%b-%Y')}"
else:
    date_display_str = "No data available"

st.markdown(f"""
<div class="dashboard-header">
    <div class="header-title-box">
        <div class="header-main-title">DAILY RISK BASED CONTRACTOR SAFETY FIELD AUDITS DASHBOARD</div>
        <div class="header-sub-title"></div>
    </div>
    <div class="header-info-box">
        <div class="header-info-lbl">Reporting Window</div>
        <div class="header-info-date">📅 {date_display_str}</div>
    </div>
</div>
""", unsafe_allow_html=True)

if df_filtered.empty:
    st.warning("⚠️ No safety observation data found for the selected date range. Please choose a broader window in the sidebar.")
else:
    # ----------------- CALCULATIONS -----------------
    
    # 1. Total Contractors
    df_valid_contractors = df_filtered[~df_filtered['Contractor name'].str.lower().str.strip().isin(ignore_contractors)]
    total_contractors = df_valid_contractors['Contractor name'].nunique()
    
    # 2. Average Severity Score
    avg_severity = df_filtered['Severity'].mean()
    
    # 3. High Severity observations (4 and 5)
    high_sev_df = df_filtered[df_filtered['Severity'].isin([4, 5])]
    total_high_sev = len(high_sev_df)
    
    # Unique audit days in selected range
    unique_audit_days = df_filtered['Date'].dt.date.nunique()
    high_sev_per_day = (total_high_sev / unique_audit_days) if unique_audit_days > 0 else 0.0
    
    # Closure Rate of High Severity
    closed_high_sev = len(high_sev_df[high_sev_df['Status'].str.lower() == 'closed'])
    high_sev_closure = (closed_high_sev / total_high_sev * 100) if total_high_sev > 0 else 0.0
    
    # Unsafe Action/Condition Ratio
    unsafe_acts = df_filtered['unsafe acts'].sum()
    unsafe_cond = df_filtered['unsafe Condes'].sum()
    unsafe_ratio = (unsafe_acts / unsafe_cond) if unsafe_cond > 0 else 0.0

    # ----------------- ALIGNED KPI ROW (PURE HTML/CSS FLEXBOX) -----------------
    
    st.markdown(f"""
    <div style="display: flex; gap: 15px; margin-bottom: 25px; align-items: stretch; width: 100%; flex-wrap: wrap;">
        <!-- Card 1 -->
        <div style="flex: 1; min-width: 200px; background-color: #f4f7fc; border-radius: 10px; padding: 20px; border-left: 5px solid #0033a0; box-shadow: 0 4px 6px rgba(0,0,0,0.02); display: flex; flex-direction: column; justify-content: center; align-items: center; text-align: center; border: 1px solid #e4ecf5;">
            <div style="font-size: 11px; font-weight: bold; color: #5a6b82; text-transform: uppercase; margin-bottom: 6px; letter-spacing: 0.5px;">Total No. of Contractors</div>
            <div style="font-size: 32px; font-weight: 800; color: #002060;">👥 {total_contractors}</div>
        </div>
        <!-- Card 2 -->
        <div style="flex: 1; min-width: 200px; background-color: #f4f7fc; border-radius: 10px; padding: 20px; border-left: 5px solid #0033a0; box-shadow: 0 4px 6px rgba(0,0,0,0.02); display: flex; flex-direction: column; justify-content: center; align-items: center; text-align: center; border: 1px solid #e4ecf5;">
            <div style="font-size: 11px; font-weight: bold; color: #5a6b82; text-transform: uppercase; margin-bottom: 6px; letter-spacing: 0.5px;">Average Manpower / Day</div>
            <div style="font-size: 32px; font-weight: 800; color: #002060;">👷 {manpower}</div>
        </div>
        <!-- Card 3 -->
        <div style="flex: 1; min-width: 200px; background-color: #f4f7fc; border-radius: 10px; padding: 20px; border-left: 5px solid #0033a0; box-shadow: 0 4px 6px rgba(0,0,0,0.02); display: flex; flex-direction: column; justify-content: center; align-items: center; text-align: center; border: 1px solid #e4ecf5;">
            <div style="font-size: 11px; font-weight: bold; color: #5a6b82; text-transform: uppercase; margin-bottom: 6px; letter-spacing: 0.5px;">Site Severity Score</div>
            <div style="font-size: 32px; font-weight: 800; color: #002060;">🛡️ {avg_severity:.1f}</div>
        </div>
        <!-- High Severity Stat Block -->
        <div style="flex: 3; min-width: 450px; background-color: #fff6f6; border-radius: 10px; padding: 15px 20px; border: 1px solid #ffe3e3; border-left: 5px solid #fa5252; box-shadow: 0 4px 6px rgba(0,0,0,0.02); display: flex; flex-direction: column; justify-content: space-between;">
            <div style="font-size: 12px; font-weight: 800; color: #fa5252; text-transform: uppercase; text-align: center; margin-bottom: 10px; letter-spacing: 0.8px;">🚨 High Severity Observations (Severity 4 & 5)</div>
            <div style="display: flex; justify-content: space-around; align-items: center; width: 100%; height: 100%;">
                <div style="text-align: center;">
                    <div style="font-size: 10px; color: #7f8c8d; text-transform: uppercase; letter-spacing: 0.5px;">Total Observations</div>
                    <div style="font-size: 26px; font-weight: 800; color: #fa5252; margin-top: 3px;">{total_high_sev}</div>
                </div>
                <div style="width: 1px; height: 35px; background-color: #ffd8d8;"></div>
                <div style="text-align: center;">
                    <div style="font-size: 10px; color: #7f8c8d; text-transform: uppercase; letter-spacing: 0.5px;">Observations / Day</div>
                    <div style="font-size: 26px; font-weight: 800; color: #fa5252; margin-top: 3px;">{high_sev_per_day:.1f}</div>
                </div>
                <div style="width: 1px; height: 35px; background-color: #ffd8d8;"></div>
                <div style="text-align: center;">
                    <div style="font-size: 10px; color: #7f8c8d; text-transform: uppercase; letter-spacing: 0.5px;">Closure Rate</div>
                    <div style="font-size: 26px; font-weight: 800; color: #40c057; margin-top: 3px;">{high_sev_closure:.0f}%</div>
                </div>
            </div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ----------------- VISUALIZATIONS SECTION (ROW 1) -----------------
    row1_c1, row1_c2, row1_c3 = st.columns([1, 1, 1.2])
    
    with row1_c1:
        with st.container(border=True):
            st.markdown("<h4 style='margin:0; padding-bottom:8px; border-bottom:2px solid #0033a0; color:#002060; font-weight:800; font-size:14px; text-transform:uppercase;'>📊 TOP OBSERVATION CATEGORIES</h4>", unsafe_allow_html=True)
            # Count trends
            df_trends = df_filtered[df_filtered['Trend'].notna() & (df_filtered['Trend'] != "") & (df_filtered['Trend'].str.lower() != "nan")]
            if not df_trends.empty:
                trend_counts = df_trends['Trend'].value_counts().reset_index()
                trend_counts.columns = ['Category', 'Count']
                trend_counts = trend_counts.sort_values('Count', ascending=True)
                
                fig_trends = px.bar(
                    trend_counts.tail(8),
                    x='Count',
                    y='Category',
                    orientation='h',
                    color_discrete_sequence=['#0033a0'],
                    text='Count'
                )
                fig_trends.update_layout(
                    margin=dict(l=10, r=10, t=10, b=10),
                    height=240,
                    xaxis_title="",
                    yaxis_title="",
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(family="Outfit, sans-serif")
                )
                st.plotly_chart(fig_trends, use_container_width=True, config={'displayModeBar': False})
            else:
                st.markdown("<div style='height:240px; display:flex; align-items:center; justify-content:center; color:#7f8c8d; font-size:14px;'>No classified Trend data in this date range.</div>", unsafe_allow_html=True)
            
    with row1_c2:
        with st.container(border=True):
            st.markdown("<h4 style='margin:0; padding-bottom:8px; border-bottom:2px solid #0033a0; color:#002060; font-weight:800; font-size:14px; text-transform:uppercase;'>🎯 ACTION / CONDITION RATIO</h4>", unsafe_allow_html=True)
            # Gauge chart for Unsafe Ratio
            fig_gauge = go.Figure(go.Indicator(
                mode = "gauge+number",
                value = unsafe_ratio,
                domain = {'x': [0, 1], 'y': [0, 1]},
                gauge = {
                    'axis': {'range': [0, 1.5], 'tickwidth': 1, 'tickcolor': "#002060"},
                    'bar': {'color': "#002060"},
                    'steps': [
                        {'range': [0, 0.4], 'color': "#d4edda"},
                        {'range': [0.4, 0.8], 'color': "#fff3cd"},
                        {'range': [0.8, 1.5], 'color': "#f8d7da"}
                    ],
                    'threshold': {
                        'line': {'color': "#fa5252", 'width': 4},
                        'thickness': 0.75,
                        'value': unsafe_ratio
                    }
                }
            ))
            fig_gauge.update_layout(
                margin=dict(l=20, r=20, t=15, b=10),
                height=210,
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(family="Outfit, sans-serif")
            )
            st.plotly_chart(fig_gauge, use_container_width=True, config={'displayModeBar': False})
            st.markdown(f"<div style='text-align:center; font-size:12px; color:#5a6b82; font-weight:600;'>Acts: <b>{unsafe_acts}</b> &nbsp;|&nbsp; Conditions: <b>{unsafe_cond}</b></div>", unsafe_allow_html=True)
        
    with row1_c3:
        with st.container(border=True):
            st.markdown("<h4 style='margin:0; padding-bottom:8px; border-bottom:2px solid #0033a0; color:#002060; font-weight:800; font-size:14px; text-transform:uppercase;'>📍 ZONE WISE AVERAGE SEVERITY</h4>", unsafe_allow_html=True)
            df_zones = df_filtered[df_filtered['Zone'].notna() & (df_filtered['Zone'] != "") & (df_filtered['Zone'].str.lower() != "nan")]
            if not df_zones.empty:
                zone_severity = df_zones.groupby('Zone')['Severity'].mean().reset_index()
                zone_severity.columns = ['Zone', 'Avg Severity']
                zone_severity = zone_severity.sort_values('Avg Severity', ascending=True)
                
                colors = []
                for val in zone_severity['Avg Severity']:
                    if val >= 3.0:
                        colors.append('#d93838')
                    elif val >= 2.5:
                        colors.append('#f5a623')
                    else:
                        colors.append('#27ae60')
                
                fig_zones = px.bar(
                    zone_severity,
                    x='Avg Severity',
                    y='Zone',
                    orientation='h',
                    text=zone_severity['Avg Severity'].apply(lambda x: f"{x:.1f}")
                )
                fig_zones.update_traces(marker_color=colors)
                fig_zones.update_layout(
                    margin=dict(l=10, r=10, t=10, b=10),
                    height=240,
                    xaxis_title="",
                    yaxis_title="",
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(family="Outfit, sans-serif")
                )
                st.plotly_chart(fig_zones, use_container_width=True, config={'displayModeBar': False})
            else:
                st.markdown("<div style='height:240px; display:flex; align-items:center; justify-content:center; color:#7f8c8d; font-size:14px;'>No Zone data in this date range.</div>", unsafe_allow_html=True)

    # ----------------- CONTRACTORS & TREND DETAILS (ROW 2) -----------------
    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
    row2_c1, row2_c2 = st.columns([1.2, 1])
    
    with row2_c1:
        with st.container(border=True):
            st.markdown("<h4 style='margin:0; padding-bottom:8px; border-bottom:2px solid #0033a0; color:#002060; font-weight:800; font-size:14px; text-transform:uppercase;'>🏗️ Contractor Wise Average Severity Score</h4>", unsafe_allow_html=True)
            if not df_valid_contractors.empty:
                contractor_severity = df_valid_contractors.groupby('Contractor name')['Severity'].mean().reset_index()
                contractor_severity.columns = ['Contractor', 'Avg Severity']
                contractor_severity = contractor_severity.sort_values('Avg Severity', ascending=True)
                
                colors_contractors = []
                for val in contractor_severity['Avg Severity']:
                    if val >= 3.0:
                        colors_contractors.append('#d93838')
                    elif val >= 2.5:
                        colors_contractors.append('#f5a623')
                    else:
                        colors_contractors.append('#27ae60')
                        
                fig_contractors = px.bar(
                    contractor_severity,
                    x='Avg Severity',
                    y='Contractor',
                    orientation='h',
                    text=contractor_severity['Avg Severity'].apply(lambda x: f"{x:.1f}")
                )
                fig_contractors.update_traces(marker_color=colors_contractors)
                fig_contractors.update_layout(
                    margin=dict(l=10, r=10, t=10, b=10),
                    height=300,
                    xaxis_title="",
                    yaxis_title="",
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(family="Outfit, sans-serif")
                )
                st.plotly_chart(fig_contractors, use_container_width=True, config={'displayModeBar': False})
            else:
                st.markdown("<div style='height:300px; display:flex; align-items:center; justify-content:center; color:#7f8c8d; font-size:14px;'>No contractor observations in this date range.</div>", unsafe_allow_html=True)
            
    with row2_c2:
        with st.container(border=True):
            st.markdown("<h4 style='margin:0; padding-bottom:8px; border-bottom:2px solid #0033a0; color:#002060; font-weight:800; font-size:14px; text-transform:uppercase;'>📈 High Severity Trend (Selected vs Prior Period)</h4>", unsafe_allow_html=True)
            
            period_days = (end_date - start_date).days
            prior_end = start_date - timedelta(days=1)
            prior_start = prior_end - timedelta(days=period_days)
            
            curr_high = df_valid_contractors[df_valid_contractors['Severity'].isin([4, 5])]
            curr_counts = curr_high['Contractor name'].value_counts().reset_index()
            curr_counts.columns = ['Contractor', 'Current Period']
            
            df_prior = df_raw[(df_raw['Date'].dt.date >= prior_start) & (df_raw['Date'].dt.date <= prior_end)]
            prior_valid_contractors = df_prior[~df_prior['Contractor name'].str.lower().str.strip().isin(ignore_contractors)]
            prior_high = prior_valid_contractors[prior_valid_contractors['Severity'].isin([4, 5])]
            prior_counts = prior_high['Contractor name'].value_counts().reset_index()
            prior_counts.columns = ['Contractor', 'Prior Period']
            
            merged_trend = pd.merge(curr_counts, prior_counts, on='Contractor', how='outer').fillna(0)
            merged_trend['Current Period'] = merged_trend['Current Period'].astype(int)
            merged_trend['Prior Period'] = merged_trend['Prior Period'].astype(int)
            merged_trend['Diff'] = merged_trend['Current Period'] - merged_trend['Prior Period']
            
            merged_trend = merged_trend.sort_values('Current Period', ascending=False)
            
            trend_rows = []
            for idx, r in merged_trend.iterrows():
                diff_val = r['Diff']
                if diff_val > 0:
                    diff_str = f"🔺 +{diff_val}"
                elif diff_val < 0:
                    diff_str = f"🔻 {diff_val}"
                else:
                    diff_str = f"➖ 0"
                    
                trend_rows.append({
                    "Contractor": r['Contractor'],
                    "High Severity Obs (Current)": r['Current Period'],
                    "Change from Prior": diff_str
                })
                
            if trend_rows:
                st.dataframe(pd.DataFrame(trend_rows), hide_index=True, use_container_width=True, height=160)
            else:
                st.markdown("<div style='height:160px; display:flex; align-items:center; justify-content:center; color:#7f8c8d; font-size:14px;'>No high severity observations found in either period.</div>", unsafe_allow_html=True)
                
            st.markdown("""
            <div style="font-size: 11px; color: #5a6b82; border-top: 1px solid #f1f3f5; padding-top: 8px; margin-top: 8px;">
                <b>Action Items:</b><br/>
                • 🔍 Conduct immediate reviews and discussions with high-risk contractor owners.<br/>
                • 🗓️ Hold daily/weekly on-ground review meetings to check and correct conditions.
            </div>
            """, unsafe_allow_html=True)

    # ----------------- DATA QUALITY HELPERS -----------------
    st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown("<h4 style='margin:0; padding-bottom:8px; border-bottom:2px solid #0033a0; color:#002060; font-weight:800; font-size:14px; text-transform:uppercase;'>📋 Pending Tasks & Data Quality Sync Helper</h4>", unsafe_allow_html=True)
        
        df_missing_trend = df_raw[df_raw['Trend'].isna() | (df_raw['Trend'] == "") | (df_raw['Trend'].str.lower() == "nan")].copy()
        
        if not df_missing_trend.empty:
            st.warning(f"⚠️ There are {len(df_missing_trend)} observations in the Excel sheet that are missing a Trend classification. Please classify them in the Excel sheet.")
            
            df_missing_display = df_missing_trend[['ln', 'Date', 'Contractor name', 'Observation Discription ', 'Severity ']].copy()
            df_missing_display['Date'] = df_missing_display['Date'].dt.strftime('%d-%b-%Y')
            df_missing_display.columns = ['Row # (ln)', 'Date', 'Contractor', 'Observation Description', 'Severity']
            
            st.dataframe(df_missing_display.head(15), hide_index=True, use_container_width=True)
            st.caption("ℹ️ Only showing the first 15 records. Please assign a category (e.g. PPE, Work at Height, Housekeeping, Electrical, etc.) in Column B of the Excel sheet.")
        else:
            st.success("✅ All sync observations have been classified! Excellent data quality.")
