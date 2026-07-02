import os
import pandas as pd
from openpyxl import load_workbook

def automate_csfa_data_entry(drcsfa_path, csfa_path, start_row, end_row, output_path="CSFA_updated.xlsx"):
    print("Loading sheets...")
    # Read Form Responses (DRCSFA)
    df_source = pd.read_excel(drcsfa_path)

    # Convert Excel row numbers (1-indexed) to pandas index positions (0-based)
    start_idx = max(0, start_row - 2)
    end_idx = min(len(df_source), end_row - 1)
    df_chunk = df_source.iloc[start_idx:end_idx + 1]

    if df_chunk.empty:
        print("Error: No rows found matching your input range bounds.")
        return

    obs_groups = [
        ('Contractor\xa0', 'Condition', 'Type of Violation', 'Severity Score', 'Violation Description', 'No of Good Citizen', 'No of Violators'),
        ('Contractor\xa02', 'Condition2', 'Type of Violation2', 'Severity Score2', 'Violation Description2', 'No of Good Citizen2', 'No of Violators2'),
        ('Contractor\xa03', 'Condition3', 'Type of Violation3', 'Severity Score3', 'Violation Description3', 'No of Good Citizen3', 'No of Violators3'),
        ('Contractor\xa04', 'Condition4', 'Type of Violation4', 'Severity Score4', 'Violation Description4', 'No of Good Citizen4', 'No of Violators4'),
        ('Contractor\xa05', 'Condition5', 'Type of Violation5', 'Severity Score5', 'Violation Description5', 'No of Good Citizen5', 'No of Violators5'),
        ('Contractor\xa06', 'Condition6', 'Type of Violation6', 'Severity Score6', 'Violation Description6', 'No of Good Citizen6', 'No of Violators6'),
        ('Contractor\xa07', 'Condition7', 'Type of Violation7', 'Severity Score7', 'Violation Description7', 'No of Good Citizen7', 'No of Violators7'),
        ('Contractor\xa08', 'Condition8', 'Type of Violation8', 'Severity Score8', 'Violation Description8', 'No of Good Citizen8', 'No of Violators8'),
        ('Contractor\xa09', 'Condition9', 'Type of Violation9', 'Severity Score9', 'Violation Description9', 'No of Good Citizen9', 'No of Violators9'),
    ]

    print(f"Processing DRCSFA Excel rows {start_row} to {end_row}...")
    new_rows = []
    for _, row in df_chunk.iterrows():
        # raw_date = row.get('Date of Audit')
        # date_str = pd.to_datetime(raw_date).strftime('%d-%m-%Y') if pd.notna(raw_date) else ""
        date_str = row.get('Date of Audit')
        zone = row.get('Zone', '')

        for group in obs_groups:
            col_contractor, col_condition, col_violation_type, col_severity, col_desc, col_good, col_violators = group
            desc_val = row.get(col_desc)

            if pd.isna(desc_val) or str(desc_val).strip() == "":
                continue

            contractor_val = row.get(col_contractor, '')
            condition_val = str(row.get(col_condition, ''))
            violation_type_val = str(row.get(col_violation_type, ''))

            raw_severity = str(row.get(col_severity, '0')).strip()
            if "-" in raw_severity:
                severity_val = raw_severity.split("-")[0].strip()
            else:
                severity_val = raw_severity

            try:
                severity_val = int(float(severity_val))
            except ValueError:
                pass

            good_val = row.get(col_good, 0)
            violators_val = row.get(col_violators, 0)

            unsafe_cond = 1 if "unsafe condition" in condition_val.lower() else ""
            unsafe_act = 1 if "unsafe act" in condition_val.lower() else ""
            ppe_non_compliance = 1 if "ppe" in violation_type_val.lower() else ""
            four_five_indicator = 1 if severity_val in [4, 5, '4', '5'] else ""
            num_violations = 1

            try:
                vx_severity = num_violations * int(severity_val)
            except:
                vx_severity = ""

            new_row = {
                'Description': str(desc_val).strip(),
                'Engineer': '',
                'Contractors': '',
                'Good Citizens': good_val if pd.notna(good_val) else 0,
                'Violators': violators_val if pd.notna(violators_val) else 0,
                'Number of Violations': num_violations,
                'Severity ': severity_val,
                'Violations x Severity': vx_severity,
                '': '',
                '4 & 5': four_five_indicator,
                'PPE Non-compliance ': ppe_non_compliance,
                'unsafe Condes': unsafe_cond,
                'unsafe acts': unsafe_act,
                'Contractor name': contractor_val,
                'Status': 'Pending',
                'Zone': zone,
                'Date': date_str
            }
            new_rows.append(new_row)

    if not new_rows:
        print("No valid data found to copy over in that row range.")
        return

    # WRITE BACK INTO EXCEL STRUCTURE
    print(f"Appending {len(new_rows)} observations to CSFA...")
    wb = load_workbook(csfa_path)
    sheet_name = 'Sheet2' if 'Sheet2' in wb.sheetnames else wb.active.title
    ws = wb[sheet_name]

    # --- FIX 1: Safely find max row with data to prevent iterating over thousands of blank rows ---
    max_populated_row = ws.max_row
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, 5)):
            max_populated_row = r
            break

    # Dynamically find next sequential serial sequence index (ln)
    start_ln = 1
    for r in range(max_populated_row, 11, -1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, (int, float)):
            start_ln = int(val) + 1
            break

    # --- FIX 2: Safely locate target row WITHOUT modifying the sheet layout mid-loop ---
    target_row = max_populated_row + 1
    for r in range(12, max_populated_row + 2):
        desc_cell = ws.cell(row=r, column=2).value
        if desc_cell is None or "sub total" in str(desc_cell).lower() or "totals" in str(desc_cell).lower():
            target_row = r
            break

    col_order = [
        'Description', 'Engineer', 'Contractors', 'Good Citizens', 'Violators',
        'Number of Violations', 'Severity ', 'Violations x Severity', '',
        '4 & 5', 'PPE Non-compliance ', 'unsafe Condes', 'unsafe acts',
        'Contractor name', 'Status', 'Zone', 'Date'
    ]

    # Insert rows strictly outside/after the scanning loops are complete
    if ws.cell(row=target_row, column=2).value is not None or ws.cell(row=target_row, column=1).value is not None:
        ws.insert_rows(target_row, amount=len(new_rows))

    for item in new_rows:
        ws.cell(row=target_row, column=1, value=start_ln)
        for col_idx, col_name in enumerate(col_order, start=2):
            ws.cell(row=target_row, column=col_idx, value=item[col_name])
        start_ln += 1
        target_row += 1

    wb.save(output_path)
    print(f"Process Complete! File generated as '{output_path}'.")

# --- CONFIGURE YOUR PARAMETERS AND RUN HERE ---
if __name__ == "__main__":

    # Adjust these row limits directly inside your Colab notebook cell every morning
    START_ROW = 206 # included
    END_ROW = 217

    automate_csfa_data_entry("DRCSFA(1-216).xlsx", "CSFA.xlsx", start_row=START_ROW, end_row=END_ROW)